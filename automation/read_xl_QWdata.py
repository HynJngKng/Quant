
import openpyxl as xl
import numpy as np
import matplotlib
import matplotlib.pyplot as plt
import datetime

class Holder(object):
    pass

def QW_Price_to_list(filepath, filename, sheet, print_companys=True):
    """
    Quantiwise로 받아온 가격데이터를 Class Holder로 리턴
    날짜, 회사명, 코드 등의 엑셀상 셀의 위치가 바뀔수 있으니 필요시마다 조금씩 수정해서 사용하면 됨
    페어트레이딩용으로 두개씩의 회사들을 매칭시켜서 출력하게 만들었음, 범용성있게 만들고 싶으면 수정 후 사용

    :param filepath: 파일위치(프로젝트위치 이후부터, 해당 프로젝트 폴더 내 data\Qw 폴더에 있으면 위치는 data\Qw)
    :param filename: A file name(확장자 포함)
    :param sheet: A sheet name
    :param print_companys: 페어의 이름 출력(boolean)
    :return: 결과값 리턴
    """

    wb = xl.load_workbook(filepath + "\\" +  filename, data_only=True)
    # ws = wb.get_sheet_by_name(sheet)                 # excel과 worksheet 열기
    ws = wb[sheet]

    len_days = ws.cell(row=7, column=1).value      # 날짜개수(A7)
    len_cmpny = ws.cell(row=7, column=2.).value    # 회사개수(B7)
    data_field = ws.cell(row=14, column=2.).value  # 주가인지 수정주가인지 등등(B14)

    cmpny_code = []
    cmpny_name = []
    sector_name = []
    days = []

    for n in range(0, len_cmpny):
        cmpny_code.append(ws.cell(row=8, column=n + 2).value)  # 회사코드리스트(B8:~)
        cmpny_name.append(ws.cell(row=9, column=n + 2).value)  # 회사명리스트  (B9:~)
        sector_name.append(ws.cell(row=13, column=n + 2).value)  # 섹터리스트  (B9:~)


    Prices = []
    Rtn = []
    TR = []
    for n in range(0, len_cmpny):
        Price_dummy_list = []
        Rtn_dummy_list = [0]
        TR_dummy_list = [0]

        for k in range(0,len_days):
        # for k in range(0,len_days-1239):
        # for k in range(0,len_days-1852):
        # for k in range(0,len_days-1605):

            if n == 0:

                day = datetime.datetime.date(ws.cell(row=k + 15, column=1).value)  # A열만(A15:~)  전구간
                # day = datetime.datetime.date(ws.cell(row=k + 1254, column=1).value)   # 15년부터
                # day = datetime.datetime.date(ws.cell(row=k + 1867, column=1).value)   # 1년간
                # day = datetime.datetime.date(ws.cell(row=k + 1620, column=1).value)     # 2년간

                days.append(day)

            Price_td = ws.cell(row=k + 15, column=n + 2).value   # Today 값 B열부터(B15:B__, C15:C__, ...)
            # Price_td = ws.cell(row=k + 1254, column=n + 2).value
            # Price_td = ws.cell(row=k + 1867, column=n + 2).value
            # Price_td = ws.cell(row=k + 1620, column=n + 2).value

            Price_yd = ws.cell(row=k + 14, column=n + 2).value   # Yesterday 값 B열부터...
            # Price_yd = ws.cell(row=k + 1253, column=n + 2).value
            # Price_yd = ws.cell(row=k + 1866, column=n + 2).value
            # Price_yd = ws.cell(row=k + 1619, column=n + 2).value

            Price_dummy_list.append(Price_td)
            if k > 0 and Price_yd != 0:
                Rtn_dummy_list.append(Price_td/Price_yd-1)
                TR_dummy_list.append((1+TR_dummy_list[k-1])*(Price_td/Price_yd)-1)
            elif Price_yd == 0:
                Rtn_dummy_list.append(0)
                TR_dummy_list.append(0)


        Prices.append(Price_dummy_list)
        Rtn.append(Rtn_dummy_list)
        TR.append(TR_dummy_list)



    result = Holder()                     # 결과값 출력 클래스

    result.company_name = cmpny_name
    result.company_code = cmpny_code
    result.sector_name = sector_name
    result.days = days
    result.Prices = Prices
    result.Return = Rtn
    result.TR = TR
    result.start_date = days[0]
    result.end_date = days[-1]
    result.data_field = data_field


    wb.close()

    if print_companys == True:
        for n in range(0, int(len_cmpny/2)):
            n2 = n * 2
            print("Pair" + str(n+1))
            print(cmpny_code[n2], '\t', cmpny_code[n2 + 1])
            print(cmpny_name[n2], '\t', cmpny_name[n2 + 1])
            print('---------------------------------')

    print("    파일명 :", '\t', filename)
    print("    시트명 :", '\t', sheet)
    print("    페어수 :", '\t', len_cmpny)
    print("    시작일 :", '\t', days[0])
    print("    종료일 :", '\t', days[-1])
    print("  관찰일수 :", '\t', len_days)
    print("관찰데이터 :", '\t', data_field)


    return result



########################   실행코드   ##########################################



if __name__ == '__main__':
    filepath = 'data'
    filename = 'Price_Data_Preferred_Holdings.xlsm'
    sheets = ["QW_Pref_Price", "QW_Hd_Price"]
    sheets = ["QW_Pref_Price"]   # 일단 테스트용으로 하나만

    Pairs_list = []

    for n in range(len(sheets)):
        Pairs_list.append(QW_Price_to_list(filepath, filename, sheets[n], False))

    print(Pairs_list[0].company_name[0])
    print(Pairs_list[0].days)
    print(Pairs_list[0].Prices[0])
    print(type(Pairs_list[0].Prices))

























